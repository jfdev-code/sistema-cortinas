# app/services/reporting.py
from typing import List, Dict, Any
import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import text
from sqlalchemy.orm import Session
import logging
from pathlib import Path
import json
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ReportGenerator:
    """
    Sistema centralizado para la generación de reportes del negocio.
    Permite crear reportes detallados en diferentes formatos y con análisis automático.
    """
    def __init__(self, db: Session):
        self.db = db
        self.report_path = Path("reports")
        self.report_path.mkdir(exist_ok=True)
    
    async def generar_reporte_rentabilidad(
        self,
        fecha_inicio: datetime,
        fecha_fin: datetime,
        formato: str = "excel"
    ) -> Path:
        """
        Genera un análisis detallado de rentabilidad por tipo de cortina y cliente.
        Incluye métricas clave como margen de ganancia, costo promedio y tendencias.
        """
        try:
            query = """
                SELECT 
                    d.nombre as diseno,
                    COUNT(c.id) as cantidad_cortinas,
                    AVG(c.costo_total) as costo_promedio,
                    AVG(c.precio_venta) as precio_promedio,
                    SUM(c.precio_venta - c.costo_total) as ganancia_total,
                    AVG((c.precio_venta - c.costo_total) / c.precio_venta * 100) as margen_porcentaje
                FROM cortinas c
                JOIN disenos d ON c.diseno_id = d.id
                WHERE c.fecha_creacion BETWEEN :fecha_inicio AND :fecha_fin
                GROUP BY d.id, d.nombre
                ORDER BY ganancia_total DESC
            """
            
            df_rentabilidad = pd.read_sql(
                query,
                self.db.bind,
                params={"fecha_inicio": fecha_inicio, "fecha_fin": fecha_fin}
            )
            
            # Calculamos métricas adicionales
            df_rentabilidad['roi'] = (
                (df_rentabilidad['ganancia_total'] / 
                (df_rentabilidad['cantidad_cortinas'] * df_rentabilidad['costo_promedio'])) * 100
            )
            
            # Agregamos análisis de tendencias mensuales
            query_tendencias = """
                SELECT 
                    DATE_TRUNC('month', c.fecha_creacion) as mes,
                    d.nombre as diseno,
                    COUNT(c.id) as ventas_mes,
                    SUM(c.precio_venta - c.costo_total) as ganancia_mes
                FROM cortinas c
                JOIN disenos d ON c.diseno_id = d.id
                WHERE c.fecha_creacion BETWEEN :fecha_inicio AND :fecha_fin
                GROUP BY DATE_TRUNC('month', c.fecha_creacion), d.id, d.nombre
                ORDER BY mes, diseno
            """
            
            df_tendencias = pd.read_sql(
                query_tendencias,
                self.db.bind,
                params={"fecha_inicio": fecha_inicio, "fecha_fin": fecha_fin}
            )
            
            # Creamos análisis de pivot para ver tendencias
            pivot_tendencias = df_tendencias.pivot_table(
                index='mes',
                columns='diseno',
                values=['ventas_mes', 'ganancia_mes'],
                aggfunc='sum'
            )
            
            # Generamos el reporte en el formato solicitado
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if formato == "excel":
                return await self._guardar_excel_multiple(
                    {
                        'Resumen Rentabilidad': df_rentabilidad,
                        'Tendencias Mensuales': df_tendencias,
                        'Análisis Pivot': pivot_tendencias
                    },
                    f"rentabilidad_{timestamp}.xlsx",
                    "Análisis de Rentabilidad"
                )
            elif formato == "csv":
                return await self._guardar_csv_multiple(
                    {
                        'rentabilidad': df_rentabilidad,
                        'tendencias': df_tendencias
                    },
                    f"rentabilidad_{timestamp}"
                )
            else:
                raise ValueError(f"Formato no soportado: {formato}")
                
        except Exception as e:
            logger.error(f"Error generando reporte de rentabilidad: {str(e)}")
            raise
    
    async def _guardar_excel_multiple(
        self,
        dataframes: Dict[str, pd.DataFrame],
        nombre_archivo: str,
        titulo: str
    ) -> Path:
        """
        Guarda múltiples DataFrames en diferentes hojas de un archivo Excel,
        aplicando formato profesional y análisis visual.
        """
        ruta_archivo = self.report_path / nombre_archivo
        
        # Creamos un writer de Excel
        writer = pd.ExcelWriter(ruta_archivo, engine='openpyxl')
        
        # Definimos estilos
        header_style = {
            'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
            'font': Font(color='FFFFFF', bold=True),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        # Escribimos cada DataFrame en su propia hoja
        for nombre_hoja, df in dataframes.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=1)
            
            worksheet = writer.sheets[nombre_hoja]
            
            # Agregamos el título
            worksheet.merge_cells('A1:H1')
            title_cell = worksheet['A1']
            title_cell.value = f"{titulo} - {nombre_hoja}"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Formato para encabezados
            for cell in worksheet[2]:
                cell.fill = header_style['fill']
                cell.font = header_style['font']
                cell.alignment = header_style['alignment']
            
            # Ajustamos anchos de columna
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            # Agregamos filtros
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Formato condicional para números
            for row in worksheet.iter_rows(min_row=3):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
        
        # Guardamos el archivo
        writer.save()
        return ruta_archivo
    
    async def _guardar_csv_multiple(
        self,
        dataframes: Dict[str, pd.DataFrame],
        base_nombre: str
    ) -> Dict[str, Path]:
        """
        Guarda múltiples DataFrames en archivos CSV separados.
        Retorna un diccionario con las rutas de los archivos generados.
        """
        rutas = {}
        for nombre, df in dataframes.items():
            ruta = self.report_path / f"{base_nombre}_{nombre}.csv"
            df.to_csv(ruta, index=False)
            rutas[nombre] = ruta
        return rutas
    
    def _calcular_dias_stock(self, stock_actual: float, consumo_diario: float) -> float:
        """
        Calcula los días estimados de stock basado en el consumo promedio.
        Retorna infinito si el consumo es 0.
        """
        if consumo_diario == 0:
            return float('inf')
        return stock_actual / consumo_diario